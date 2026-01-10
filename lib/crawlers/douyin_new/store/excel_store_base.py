# -*- coding: utf-8 -*-
# Excel Store Base Implementation

import threading
from datetime import datetime
from typing import Dict, List, Any
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from base.base_crawler import AbstractStore
from tools import utils


class ExcelStoreBase(AbstractStore):
    """
    Base class for Excel storage implementation
    """

    _instances: Dict[str, "ExcelStoreBase"] = {}
    _lock = threading.Lock()

    @classmethod
    def get_instance(cls, platform: str, crawler_type: str) -> "ExcelStoreBase":
        key = f"{platform}_{crawler_type}"
        with cls._lock:
            if key not in cls._instances:
                cls._instances[key] = cls(platform, crawler_type)
            return cls._instances[key]

    @classmethod
    def flush_all(cls):
        with cls._lock:
            for key, instance in cls._instances.items():
                try:
                    instance.flush()
                    utils.logger.info(f"[ExcelStoreBase] Flushed instance: {key}")
                except Exception as e:
                    utils.logger.error(f"[ExcelStoreBase] Error flushing {key}: {e}")
            cls._instances.clear()

    def __init__(self, platform: str, crawler_type: str = "search"):
        if not EXCEL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        super().__init__()
        self.platform = platform
        self.crawler_type = crawler_type

        self.data_dir = Path("data") / platform / "excel"
        self.data_dir.mkdir(parents=True, exist_ok=True)

        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)

        self.contents_sheet = self.workbook.create_sheet("Contents")
        self.comments_sheet = self.workbook.create_sheet("Comments")
        self.creators_sheet = self.workbook.create_sheet("Creators")

        self.contents_headers_written = False
        self.comments_headers_written = False
        self.creators_headers_written = False

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filename = self.data_dir / f"{platform}_{crawler_type}_{timestamp}.xlsx"

        utils.logger.info(f"[ExcelStoreBase] Initialized Excel export to: {self.filename}")

    def _apply_header_style(self, sheet, row_num: int = 1):
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in sheet[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

    def _auto_adjust_column_width(self, sheet):
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass

            adjusted_width = min(max(max_length + 2, 10), 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _write_headers(self, sheet, headers: List[str]):
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)
        self._apply_header_style(sheet)

    def _write_row(self, sheet, data: Dict[str, Any], headers: List[str]):
        row_num = sheet.max_row + 1

        for col_num, header in enumerate(headers, 1):
            value = data.get(header, "")

            if isinstance(value, (list, dict)):
                value = str(value)
            elif value is None:
                value = ""

            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    async def store_content(self, content_item: Dict):
        headers = list(content_item.keys())
        if not self.contents_headers_written:
            self._write_headers(self.contents_sheet, headers)
            self.contents_headers_written = True
        self._write_row(self.contents_sheet, content_item, headers)
        content_id = content_item.get('aweme_id') or content_item.get('video_id') or 'N/A'
        utils.logger.info(f"[ExcelStoreBase] Stored content to Excel: {content_id}")

    async def store_comment(self, comment_item: Dict):
        headers = list(comment_item.keys())
        if not self.comments_headers_written:
            self._write_headers(self.comments_sheet, headers)
            self.comments_headers_written = True
        self._write_row(self.comments_sheet, comment_item, headers)
        utils.logger.info(f"[ExcelStoreBase] Stored comment to Excel: {comment_item.get('comment_id', 'N/A')}")

    async def store_creator(self, creator: Dict):
        headers = list(creator.keys())
        if not self.creators_headers_written:
            self._write_headers(self.creators_sheet, headers)
            self.creators_headers_written = True
        self._write_row(self.creators_sheet, creator, headers)
        utils.logger.info(f"[ExcelStoreBase] Stored creator to Excel: {creator.get('user_id', 'N/A')}")

    def flush(self):
        try:
            self._auto_adjust_column_width(self.contents_sheet)
            self._auto_adjust_column_width(self.comments_sheet)
            self._auto_adjust_column_width(self.creators_sheet)

            if self.contents_sheet.max_row == 1:
                self.workbook.remove(self.contents_sheet)
            if self.comments_sheet.max_row == 1:
                self.workbook.remove(self.comments_sheet)
            if self.creators_sheet.max_row == 1:
                self.workbook.remove(self.creators_sheet)

            if len(self.workbook.sheetnames) == 0:
                utils.logger.info(f"[ExcelStoreBase] No data to save, skipping file creation: {self.filename}")
                return

            self.workbook.save(self.filename)
            utils.logger.info(f"[ExcelStoreBase] Excel file saved successfully: {self.filename}")

        except Exception as e:
            utils.logger.error(f"[ExcelStoreBase] Error saving Excel file: {e}")
            raise
